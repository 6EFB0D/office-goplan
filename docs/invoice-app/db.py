# -*- coding: utf-8 -*-
"""Excel ベースの取引管理データベース操作モジュール。"""
from __future__ import annotations

import datetime
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config

# ---- スキーマ ----

HEADERS = [
    "取引ID", "作成日", "顧客名", "顧客住所",
    "品目", "数量", "単価", "金額",
    "見積番号", "見積日",
    "請求番号", "請求日", "入金期限",
    "領収番号", "領収日",
    "ステータス", "備考",
]
COL = {h: i + 1 for i, h in enumerate(HEADERS)}  # 1-indexed

STATUS = {
    "new":       "新規",
    "quoted":    "見積発行済",
    "invoiced":  "請求発行済",
    "paid":      "入金確認済",
    "completed": "完了",
}

_COL_WIDTHS = {
    "取引ID": 20, "作成日": 12, "顧客名": 22, "顧客住所": 36,
    "品目": 36, "数量": 6, "単価": 9, "金額": 10,
    "見積番号": 22, "見積日": 12,
    "請求番号": 22, "請求日": 12, "入金期限": 12,
    "領収番号": 22, "領収日": 12,
    "ステータス": 14, "備考": 26,
}

# ---- ワークブック初期化 ----

def _create_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "取引管理"
    hfont  = Font(bold=True, color="FFFFFF", name="Meiryo UI", size=10)
    hfill  = PatternFill("solid", fgColor="1A2332")
    halign = Alignment(horizontal="center", vertical="center")
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(1, col, h)
        c.font, c.fill, c.alignment = hfont, hfill, halign
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    for h, w in _COL_WIDTHS.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(COL[h])].width = w

    ws2 = wb.create_sheet("採番")
    ws2.append(["年", "Q最終連番", "I最終連番", "R最終連番"])
    ws2.append([datetime.date.today().year, 0, 0, 0])
    for c in ws2[1]:
        c.font = Font(bold=True, name="Meiryo UI")
    for col, w in enumerate([8, 14, 14, 14], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    os.makedirs(os.path.dirname(config.EXCEL_PATH), exist_ok=True)
    wb.save(config.EXCEL_PATH)
    return wb


def _load() -> openpyxl.Workbook:
    if not os.path.exists(config.EXCEL_PATH):
        return _create_workbook()
    return openpyxl.load_workbook(config.EXCEL_PATH)


# ---- 採番 ----

def next_number(prefix: str) -> str:
    """次の書類番号を返し、採番シートを更新する。例: Q-20260425-0001"""
    wb = _load()
    ws = wb["採番"]
    today = datetime.date.today()
    col_map = {"Q": 2, "I": 3, "R": 4}
    col = col_map[prefix]

    target = None
    for row in ws.iter_rows(min_row=2):
        if row[0].value == today.year:
            target = row
            break
    if target is None:
        nr = ws.max_row + 1
        ws.cell(nr, 1, today.year)
        ws.cell(nr, 2, 0); ws.cell(nr, 3, 0); ws.cell(nr, 4, 0)
        target = [ws.cell(nr, c) for c in range(1, 5)]

    current = target[col - 1].value or 0
    nxt = current + 1
    target[col - 1].value = nxt
    wb.save(config.EXCEL_PATH)
    return f"{prefix}-{today.strftime('%Y%m%d')}-{nxt:04d}"


def _next_txn_id(ws) -> str:
    today = datetime.date.today()
    seq = ws.max_row  # header=1, first data → seq=2 → TRX-...-002
    return f"TRX-{today.strftime('%Y%m%d')}-{seq:03d}"


# ---- CRUD ----

def create_transaction(customer_name: str, customer_address: str,
                       items: list[dict], notes: str = "") -> str:
    """
    items: [{"description": str, "qty": int, "unit_price": int}]
    戻り値: 取引ID
    """
    wb = _load()
    ws = wb["取引管理"]

    item   = items[0]
    total  = sum(i["qty"] * i["unit_price"] for i in items)
    txn_id = _next_txn_id(ws)
    today  = datetime.date.today()

    row = [None] * len(HEADERS)
    row[COL["取引ID"]    - 1] = txn_id
    row[COL["作成日"]    - 1] = today
    row[COL["顧客名"]    - 1] = customer_name
    row[COL["顧客住所"]  - 1] = customer_address
    row[COL["品目"]      - 1] = item["description"]
    row[COL["数量"]      - 1] = item["qty"]
    row[COL["単価"]      - 1] = item["unit_price"]
    row[COL["金額"]      - 1] = total
    row[COL["ステータス"]- 1] = STATUS["new"]
    row[COL["備考"]      - 1] = notes
    ws.append(row)

    dr = ws.max_row
    alt_fill = PatternFill("solid", fgColor="F8FAFC") if dr % 2 == 0 else None
    border   = Border(bottom=Side(style="thin", color="E2E8F0"))
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(dr, c)
        cell.alignment = Alignment(vertical="center")
        cell.border = border
        if alt_fill:
            cell.fill = alt_fill
        if HEADERS[c - 1] in ("単価", "金額"):
            cell.number_format = "#,##0"
    ws.row_dimensions[dr].height = 18

    wb.save(config.EXCEL_PATH)
    return txn_id


def get_all() -> list[dict]:
    wb = _load()
    ws = wb["取引管理"]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            result.append(dict(zip(HEADERS, row)))
    return result


def get_by_id(txn_id: str) -> dict | None:
    for t in get_all():
        if t.get("取引ID") == txn_id:
            return t
    return None


def update(txn_id: str, **fields) -> None:
    """指定した取引の任意フィールドを更新する。"""
    wb = _load()
    ws = wb["取引管理"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == txn_id:
            for key, val in fields.items():
                if key in COL:
                    ws.cell(row[0].row, COL[key], val)
            break
    wb.save(config.EXCEL_PATH)
